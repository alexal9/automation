import openpyxl
from openpyxl.utils import get_column_letter, coordinate_from_string
from collections import defaultdict

wb = openpyxl.load_workbook("translations.xlsx")

result = defaultdict(dict)

# extract translation info
for name in wb.get_sheet_names():
	sheet = wb.get_sheet_by_name(name)
	rows = sheet.max_row
	cols = sheet.max_column
	end = get_column_letter(cols) + str(rows)
	for rowOfCellObj in sheet['C2':end]:
		for cellObj in rowOfCellObj:
			col, row = coordinate_from_string(cellObj.coordinate)
			if cellObj.value:
				result[ sheet['B'+str(row)].value] [sheet[col+'1'].value] = cellObj.value

text = [i for i in result]
languages = [ i for i in result[text[0]] ]

# write translation mapping to another excel file

wbout = openpyxl.Workbook()
wbout.remove_sheet(wbout.get_sheet_by_name('Sheet'))
for num, language in enumerate(sorted(languages)):
	wbout.create_sheet(index = num, title = language)

for language in languages:
	sheet = wbout.get_sheet_by_name(language)
	for row, line in enumerate(sorted(text)):
		sheet['A'+str(row+1)] = line
		try:
			sheet['B'+str(row+1)] = result[line][language]
		except:
			print('missing text:', line, 'for language', language)

wbout.save('remapped.xlsx')
